import os
import pandas as pd
import json
from flask import Flask, jsonify, make_response, render_template, request, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime

# Initialize the Flask app
app = Flask(__name__)

# Define the directory for the CSV data files
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')


def parse_evidence_json(evidence_str):
    """Parse Evidence JSON string and extract image URLs."""
    try:
        if not evidence_str or evidence_str.strip() == '':
            return []
        
        evidence_data = json.loads(evidence_str)
        if isinstance(evidence_data, list):
            return [item.get('url') for item in evidence_data if item.get('type') == 'image' and item.get('url')]
        return []
    except (json.JSONDecodeError, AttributeError, TypeError):
        return []

def get_priority_color(priority):
    """Get color fill for priority/severity levels."""
    if not priority:
        return None
    
    priority_lower = str(priority).lower().strip()
    
    if 'critical' in priority_lower:
        return PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark red
    elif 'high' in priority_lower:
        return PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")  # Orange red
    elif 'medium' in priority_lower:
        return PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    elif 'low' in priority_lower:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    else:
        return None

def autofit_worksheet(ws):
    """Auto-fit column widths and row heights for the worksheet."""
    # Auto-fit column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    # Calculate length considering line breaks
                    cell_value = str(cell.value)
                    lines = cell_value.split('\n')
                    max_line_length = max(len(line) for line in lines)
                    max_length = max(max_length, max_line_length)
            except:
                pass
        
        # Set column width (minimum 15, maximum 60 for better image display)
        adjusted_width = min(max(max_length + 2, 15), 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Auto-fit row heights with better image support
    for row in ws.iter_rows():
        max_height = 20  # Increased minimum row height
        
        for cell in row:
            if cell.value:
                try:
                    cell_value = str(cell.value)
                    lines = cell_value.split('\n')
                    line_count = len(lines)
                    
                    # Calculate height based on number of lines (20 points per line for better spacing)
                    calculated_height = line_count * 20
                    max_height = max(max_height, calculated_height)
                    
                    # Check if this is an Evidence cell with images
                    if "📷" in str(cell.value) and "image(s)" in str(cell.value):
                        # Add extra height for images (minimum 100 points for image cells)
                        max_height = max(max_height, 100)
                except:
                    pass
        
        # Set row height (minimum 20, maximum 400 for better image display)
        adjusted_height = min(max(max_height, 20), 400)
        ws.row_dimensions[row[0].row].height = adjusted_height

@app.route('/')
def index():
    """Serves the main index.html page."""
    return render_template('index.html')

@app.route('/data/<path:filename>')
def serve_data_file(filename):
    """Serves a file from the 'data' directory."""
    response = make_response(send_from_directory(DATA_DIR, filename))
    response.headers['Content-Type'] = 'text/csv'
    return response

@app.route('/save-data', methods=['POST'])
def save_data():
    """Saves the provided CSV data to the file."""
    try:
        data = request.get_json()
        csv_data = data.get('data')
        
        if csv_data is None:
            return jsonify(success=False, error="No data provided"), 400

        file_path = os.path.join(DATA_DIR, '.test.csv')
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            f.write(csv_data)
            
        return jsonify(success=True)
    except Exception as e:
        print(f"Error saving data: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route('/export-xlsx', methods=['POST'])
def export_xlsx():
    """Exports data to XLSX with styled sheets for bug reports, test cases, and analysis."""
    try:
        data = request.get_json()
        csv_data = data.get('data')
        
        if csv_data is None:
            return jsonify(success=False, error="No data provided"), 400

        # Create DataFrame from CSV data
        from io import StringIO
        df = pd.read_csv(StringIO(csv_data))
        print(f"Successfully created DataFrame with {len(df)} rows")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Filter data by type
        test_cases = df[df['Type'] == 'Test Case'].copy()
        bug_reports = df[df['Type'] == 'Bug Report'].copy()
        print(f"Found {len(test_cases)} test cases and {len(bug_reports)} bug reports")
        
        # 1. Test Cases Sheet
        if not test_cases.empty:
            ws_tc = wb.create_sheet("Test Cases")
            ws_tc.title = "Test Cases"
            
            # Add headers
            headers = ['ID', 'Platform', 'Section', 'Title/Objective', 'Test Steps', 'Test Data', 
                      'Expected Result', 'Actual Result', 'Priority/Severity', 'Status', 'Description']
            
            for col, header in enumerate(headers, 1):
                cell = ws_tc.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Add data
            for row_idx, (_, row) in enumerate(test_cases.iterrows(), 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_tc.cell(row=row_idx, column=col_idx, value=row.get(header, ''))
                    cell.border = border
                    # Enable text wrapping for better display
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Priority/Severity column coloring
                    if col_idx == 9 and header == 'Priority/Severity':  # Priority/Severity column
                        priority_color = get_priority_color(row.get('Priority/Severity', ''))
                        if priority_color:
                            cell.fill = priority_color
                    
                    # Status column coloring
                    if col_idx == 10:  # Status column
                        if str(row.get('Status', '')).upper() == 'PASS':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif str(row.get('Status', '')).upper() == 'FAIL':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Auto-fit the worksheet
            autofit_worksheet(ws_tc)
        
        # 2. Bug Reports Sheet
        if not bug_reports.empty:
            ws_br = wb.create_sheet("Bug Reports")
            ws_br.title = "Bug Reports"
            
            # Add headers
            headers = ['ID', 'Platform', 'Section', 'Title/Objective', 'Test Steps', 'Test Data', 
                      'Expected Result', 'Actual Result', 'Priority/Severity', 'Status', 'Description', 'Evidence']
            
            for col, header in enumerate(headers, 1):
                cell = ws_br.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                cell.border = border
                cell.alignment = center_alignment
            
            # Add data
            for row_idx, (_, row) in enumerate(bug_reports.iterrows(), 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_br.cell(row=row_idx, column=col_idx, value=row.get(header, ''))
                    cell.border = border
                    # Enable text wrapping for better display
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Priority/Severity column coloring
                    if col_idx == 9 and header == 'Priority/Severity':  # Priority/Severity column
                        priority_color = get_priority_color(row.get('Priority/Severity', ''))
                        if priority_color:
                            cell.fill = priority_color
                    
                    # Status column coloring
                    if col_idx == 10:  # Status column
                        if str(row.get('Status', '')).upper() == 'OPEN':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif str(row.get('Status', '')).upper() == 'CLOSED':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif str(row.get('Status', '')).upper() == 'IN PROGRESS':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
                    # Handle Evidence column with hyperlinks
                    if col_idx == 12 and header == 'Evidence':  # Evidence column
                        evidence_str = str(row.get('Evidence', ''))
                        if evidence_str and evidence_str != 'nan':
                            # Parse evidence JSON to get image URLs
                            image_urls = parse_evidence_json(evidence_str)
                            
                            if image_urls:
                                try:
                                    # Create hyperlinks for each image URL
                                    hyperlink_text = f"📷 {len(image_urls)} image(s):\n\n"
                                    for i, url in enumerate(image_urls, 1):
                                        hyperlink_text += f"Image {i}: {url}\n"
                                    
                                    cell.value = hyperlink_text
                                    
                                    # Add hyperlinks to the cell
                                    if image_urls:
                                        cell.hyperlink = image_urls[0]
                                        cell.font = Font(color="0000FF", underline="single")
                                    
                                    # Set appropriate row height for multiple URLs
                                    line_count = len(image_urls) + 2  # +2 for header and spacing
                                    ws_br.row_dimensions[row_idx].height = max(line_count * 15, 60)
                                except Exception as e:
                                    print(f"Error creating hyperlink for row {row_idx}: {e}")
                                    cell.value = f"📷 {len(image_urls)} image(s) - URLs: {', '.join(image_urls)}"
                            else:
                                cell.value = evidence_str
                        else:
                            cell.value = evidence_str
            
            # Auto-fit the worksheet
            autofit_worksheet(ws_br)
        
        # 3. Analysis Sheet
        ws_analysis = wb.create_sheet("Analysis")
        ws_analysis.title = "Analysis"
        
        # Add summary statistics
        analysis_data = [
            ['Metric', 'Count', 'Percentage'],
            ['Total Test Cases', len(test_cases), f"{len(test_cases)/len(df)*100:.1f}%" if len(df) > 0 else "0%"],
            ['Total Bug Reports', len(bug_reports), f"{len(bug_reports)/len(df)*100:.1f}%" if len(df) > 0 else "0%"],
            ['', '', ''],
            ['Test Case Status', '', ''],
            ['Passed', len(test_cases[test_cases['Status'].str.upper() == 'PASS']) if not test_cases.empty else 0, ''],
            ['Failed', len(test_cases[test_cases['Status'].str.upper() == 'FAIL']) if not test_cases.empty else 0, ''],
            ['', '', ''],
            ['Bug Report Status', '', ''],
            ['Open', len(bug_reports[bug_reports['Status'].str.upper() == 'OPEN']) if not bug_reports.empty else 0, ''],
            ['Closed', len(bug_reports[bug_reports['Status'].str.upper() == 'CLOSED']) if not bug_reports.empty else 0, ''],
            ['In Progress', len(bug_reports[bug_reports['Status'].str.upper() == 'IN PROGRESS']) if not bug_reports.empty else 0, ''],
            ['', '', ''],
            ['Platform Distribution', '', ''],
        ]
        
        # Add platform distribution
        if not df.empty:
            platform_counts = df['Platform'].value_counts()
            for platform, count in platform_counts.items():
                analysis_data.append([platform, count, f"{count/len(df)*100:.1f}%"])
        
        # Add analysis data to sheet
        for row_idx, row_data in enumerate(analysis_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_analysis.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.alignment = center_alignment
                elif row_idx in [2, 3, 5, 6, 9, 10, 11, 13]:  # Data rows
                    cell.alignment = center_alignment
                    if row_idx in [2, 3]:  # Total counts
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif row_idx in [6, 7]:  # Test case status
                        if 'Passed' in str(value):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif 'Failed' in str(value):
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif row_idx in [10, 11, 12]:  # Bug report status
                        if 'Open' in str(value):
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif 'Closed' in str(value):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif 'In Progress' in str(value):
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Auto-fit the analysis worksheet
        autofit_worksheet(ws_analysis)
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        print("Saving workbook to BytesIO...")
        wb.save(output)
        output.seek(0)
        print("Workbook saved successfully")
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=testing_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        print("XLSX export completed successfully")
        return response
        
    except Exception as e:
        print(f"Error exporting XLSX: {e}")
        return jsonify(success=False, error=str(e)), 500

if __name__ == '__main__':
    # Runs the app on a local development server
    app.run(debug=True)